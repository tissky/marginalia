"""Spreadsheet pipeline (.xlsx / .xlsm via openpyxl).

Renders sheets as a markdown-ish text view: one `# Sheet: <name>` heading
per sheet, then up to N rows. Long sheets are sampled — the first M rows
plus a tail summary — so the LLM gets a sense of structure without
swallowing a 100k-row spreadsheet.

read_segment supports `heading="Sheet: <name>"` to fetch one sheet's
rendered rows, `pattern` for regex search, and the generic
offset/max_chars chunking. For tabular querying use `query_sql`.

XLS (legacy binary) is not supported; users should resave to .xlsx.
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any

from marginalia.pipelines._text_indexer import index_extracted_text
from marginalia.pipelines.base import (
    Pipeline,
    PipelineContext,
    PipelineResult,
    SegmentResult,
)
from marginalia.pipelines.registry import register_pipeline
from marginalia.storage.base import StorageBackend

log = logging.getLogger(__name__)

MAX_XLSX_BYTES = 30 * 1024 * 1024
MAX_ROWS_PER_SHEET = 200
MAX_TAIL_PEEK = 20
MAX_CELL_CHARS = 200
DEFAULT_MAX_CHARS = 8000


@register_pipeline(
    mimes=(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    ),
    exts=(".xlsx", ".xlsm"),
)
class SpreadsheetPipeline(Pipeline):
    name = "spreadsheet"

    async def run(
        self,
        *,
        ctx: PipelineContext,
        storage: StorageBackend,
    ) -> PipelineResult:
        body = await self._extract_text(storage, ctx.storage_key)
        return await index_extracted_text(body, ctx, kind="spreadsheet")

    async def read_segment(
        self,
        *,
        file_row: Any,
        args: dict[str, Any],
        storage: StorageBackend,
    ) -> SegmentResult:
        """Resolve args against the rendered workbook.

        Field priority:
          1. pattern              → regex search over rendered text
          2. heading              → "Sheet: <name>" → that sheet's rows
          3. (default)            → offset..offset+max_chars chunk
        """
        body = await self._extract_text(storage, file_row.storage_key)

        offset = max(0, int(args.get("offset") or 0))
        max_chars = int(args.get("max_chars") or DEFAULT_MAX_CHARS)
        if max_chars <= 0:
            max_chars = DEFAULT_MAX_CHARS

        pattern = (args.get("pattern") or "").strip()
        if pattern:
            return _ss_pattern_search(
                body=body, pattern=pattern,
                context_lines=int(args.get("context_lines") or 2),
                max_matches=int(args.get("max_matches") or 20),
            )

        heading = (args.get("heading") or "").strip()
        if heading:
            slab = _slice_by_heading(body, heading)
            if slab is None:
                sheet_names = _list_sheet_headings(body)
                return SegmentResult(
                    error=f"sheet/heading not found: {heading!r}",
                    extras={"available_sheets": sheet_names},
                )
            return _clamp_ss(
                slab, offset, max_chars,
                extras={"heading": heading},
            )

        return _clamp_ss(body, offset, max_chars)

    @staticmethod
    async def _extract_text(storage: StorageBackend, key: str) -> str:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "spreadsheet pipeline needs openpyxl; "
                "`pip install openpyxl`"
            ) from exc

        buf = bytearray()
        async for chunk in storage.get(key):
            buf.extend(chunk)
            if len(buf) > MAX_XLSX_BYTES:
                raise ValueError(
                    f"xlsx exceeds {MAX_XLSX_BYTES // (1024*1024)}MB cap"
                )

        wb = openpyxl.load_workbook(
            io.BytesIO(bytes(buf)),
            data_only=True,
            read_only=True,
        )
        try:
            return _render_workbook(wb)
        finally:
            wb.close()


def _render_workbook(wb: Any) -> str:
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"# Sheet: {sheet_name}")
        rows = list(_iter_rows(ws, MAX_ROWS_PER_SHEET + MAX_TAIL_PEEK))
        if not rows:
            parts.append("(empty sheet)")
            continue
        if len(rows) <= MAX_ROWS_PER_SHEET:
            for r in rows:
                parts.append(_format_row(r))
        else:
            for r in rows[:MAX_ROWS_PER_SHEET]:
                parts.append(_format_row(r))
            parts.append(
                f"\n[…{len(rows) - MAX_ROWS_PER_SHEET}+ more rows omitted…]"
            )
        parts.append("")
    return "\n".join(parts).strip()


def _iter_rows(ws: Any, hard_limit: int):
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            yield row
            count += 1
            if count >= hard_limit:
                return


def _format_row(row: tuple) -> str:
    cells: list[str] = []
    for c in row:
        if c is None:
            cells.append("")
            continue
        s = str(c)
        if len(s) > MAX_CELL_CHARS:
            s = s[:MAX_CELL_CHARS] + "…"
        cells.append(s.replace("|", r"\|").replace("\n", " "))
    return " | ".join(cells)


# ---- read_segment helpers --------------------------------------------------

_SHEET_HEADING_RX = re.compile(r"^# Sheet: (.+)$", re.MULTILINE)


def _list_sheet_headings(body: str) -> list[str]:
    return _SHEET_HEADING_RX.findall(body)


def _slice_by_heading(body: str, heading: str) -> str | None:
    """Return the body of the named sheet (everything from its `# Sheet: name`
    line up to the next `# Sheet:` line or EOF). The heading argument may
    be the full `Sheet: name` form or just `name`."""
    target = heading.removeprefix("Sheet: ").strip()
    matches = list(_SHEET_HEADING_RX.finditer(body))
    for i, m in enumerate(matches):
        if m.group(1).strip() == target:
            start = m.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
            return body[start:end].rstrip()
    return None


def _clamp_ss(
    text: str, offset: int, max_chars: int,
    *, extras: dict[str, Any] | None = None,
) -> SegmentResult:
    extras = dict(extras or {})
    total = len(text)
    chunk = text[offset: offset + max_chars]
    truncated = (offset + len(chunk)) < total
    extras.update({
        "offset": offset,
        "char_count": len(chunk),
        "total_chars": total,
        "truncated": truncated,
        "available_sheets": _list_sheet_headings(text),
    })
    if truncated:
        extras["next_offset"] = offset + len(chunk)
    if not chunk:
        return SegmentResult(text="", error="empty result", extras=extras)
    return SegmentResult(text=chunk, extras=extras)


def _ss_pattern_search(
    *, body: str, pattern: str, context_lines: int, max_matches: int,
) -> SegmentResult:
    try:
        rx = re.compile(pattern, re.IGNORECASE | re.MULTILINE)
    except re.error as exc:
        return SegmentResult(error=f"invalid regex: {exc}")

    lines = body.splitlines()
    hits: list[dict[str, Any]] = []
    current_sheet = ""
    for i, line in enumerate(lines, start=1):
        m_sheet = _SHEET_HEADING_RX.match(line)
        if m_sheet:
            current_sheet = m_sheet.group(1).strip()
            continue
        for m in rx.finditer(line):
            if len(hits) >= max_matches:
                break
            s = max(0, i - 1 - context_lines)
            e = min(len(lines), i + context_lines)
            hits.append({
                "sheet": current_sheet,
                "line": i,
                "match": m.group(0)[:200],
                "context": "\n".join(lines[s:e]),
            })
        if len(hits) >= max_matches:
            break

    if not hits:
        return SegmentResult(
            text="", error="no matches",
            extras={
                "pattern": pattern,
                "available_sheets": _list_sheet_headings(body),
            },
        )

    rendered = "\n\n".join(
        f"[{h['sheet']} L{h['line']}] {h['match']}\n  ┊ {h['context']}"
        for h in hits
    )
    return SegmentResult(
        text=rendered,
        extras={
            "pattern": pattern,
            "match_count": len(hits),
            "hits": hits,
            "available_sheets": _list_sheet_headings(body),
        },
    )
